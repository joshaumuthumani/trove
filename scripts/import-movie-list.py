#!/usr/bin/env python3
"""Trove — one-off importer: db/source/Josh's Movie List.xlsx -> db/seed/{movies,tv}.json.

Replaces the prototype sample for Movies + TV with the owner's real catalog.
Games are left untouched (db/seed/games.json is not regenerated here).

Source shape (two sheets):
  * "Movies":   #, Name, Watched, Blu-Ray, DVD, Movies Anywhere, iTunes, YouTube,
                Amazon Video, VUDU, Microsoft Movies   (Yes = owned on that locker)
  * "TV Series": #, Series Name, Season, Episodes, iTunes, Google Play, Amazon Video,
                Movies Anywhere, VUDU, Fandango Now, XBox Video

Mapping decisions (see CLAUDE.md / src/lib/platforms.ts for the canonical vocab):
  * iTunes  -> "Apple TV"          (Apple rebrand of the iTunes store)
  * VUDU    -> "Fandango at Home"  (VUDU rebrand)
  * Microsoft Movies -> DROPPED    (store is defunct; those titles now sit under
                                    Movies Anywhere, which the sheet already marks)
  * Google Play / Movies Anywhere / Fandango Now columns on the TV sheet are
    entirely empty in the source, so they contribute nothing.
  * No TMDB ids / posters / years exist in the sheet, so every imported movie and
    series is flagged needs_review=1 (schema intent: "bulk backfill -> fix later").

TV "Episodes" cell:
  * "All"            -> "all"
  * a number (e.g. 1)-> [1]            (specific episode owned, e.g. a free pilot)
  * "9,11"           -> [9, 11]
  * a date (Sheets mis-parsed "1,13" as Jan-13) -> [month, day]
"""
import datetime
import json
import os

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(ROOT, "db/source/Josh's Movie List.xlsx")
OUT = os.path.join(ROOT, "db/seed")

# sheet column header -> canonical app vocabulary (None = drop)
MOVIE_DIGITAL = {
    "iTunes": "Apple TV",
    "Movies Anywhere": "Movies Anywhere",
    "Amazon Video": "Amazon Video",
    "VUDU": "Fandango at Home",
    "YouTube": "YouTube",
    "Microsoft Movies": None,  # defunct store -> covered by Movies Anywhere
}
# canonical display order (matches src/lib/platforms.ts MOVIE_DIGITAL)
DIGITAL_ORDER = ["Apple TV", "Movies Anywhere", "Amazon Video", "Fandango at Home", "YouTube"]
MOVIE_PHYSICAL = {"Blu-Ray": "Blu-Ray", "DVD": "DVD"}
PHYSICAL_ORDER = ["Blu-Ray", "DVD"]

TV_SERVICE = {
    "iTunes": "Apple TV",
    "Amazon Video": "Amazon Video",
    "XBox Video": None,  # discontinued store -> dropped (cf. Microsoft Movies)
    # empty in source, mapped defensively / dropped:
    "Google Play": None,
    "Movies Anywhere": None,
    "VUDU": None,
    "Fandango Now": None,
}
TV_ORDER = ["Apple TV", "Amazon Video"]


def is_yes(v):
    return v is not None and str(v).strip().lower() == "yes"


def clean_title(v):
    """Numeric titles (65, 300, 1917) come in as floats -> render as ints."""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    if isinstance(v, int):
        return str(v)
    return str(v).strip()


def header_map(ws):
    hdr = [c.value for c in ws[1]]
    return {h: i for i, h in enumerate(hdr) if h is not None}


def data_rows(ws, ncols):
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, ncols + 1)]
        if all(v is None for v in vals):
            continue
        yield vals


def parse_episodes(v):
    if v is None:
        return "unowned"
    if isinstance(v, (datetime.datetime, datetime.date)):
        # Sheets mis-parsed "month,day" as a date
        return [v.month, v.day]
    if isinstance(v, (int, float)):
        return [int(v)]
    s = str(v).strip()
    if s.lower() == "all":
        return "all"
    if not s:
        return "unowned"
    parts = [p.strip() for p in s.replace(";", ",").split(",") if p.strip()]
    out = []
    for p in parts:
        try:
            out.append(int(float(p)))
        except ValueError:
            pass
    return out if out else "unowned"


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)

    # ---- MOVIES ----------------------------------------------------------
    ws = wb["Movies"]
    hm = header_map(ws)
    ncols = max(hm.values()) + 1
    movies = []
    for i, row in enumerate(data_rows(ws, ncols), start=1):
        name = row[hm["Name"]]
        if name is None:
            continue
        digital_set = {
            MOVIE_DIGITAL[col]
            for col in MOVIE_DIGITAL
            if MOVIE_DIGITAL[col] and col in hm and is_yes(row[hm[col]])
        }
        physical_set = {
            MOVIE_PHYSICAL[col]
            for col in MOVIE_PHYSICAL
            if col in hm and is_yes(row[hm[col]])
        }
        movies.append({
            "id": i,
            "tmdb_id": None,
            "title": clean_title(name),
            "year": None,
            "poster_url": None,
            "digital": [s for s in DIGITAL_ORDER if s in digital_set],
            "physical": [s for s in PHYSICAL_ORDER if s in physical_set],
            "needs_review": 1,
        })

    # ---- TV (group rows by series; each row is a season) -----------------
    ws = wb["TV Series"]
    hm = header_map(ws)
    ncols = max(hm.values()) + 1
    series_order = []          # preserve first-seen order
    series_by_name = {}        # name -> series dict
    seasons = []
    for row in data_rows(ws, ncols):
        raw_name = row[hm["Series Name"]]
        if raw_name is None:
            continue
        name = clean_title(raw_name)
        if name not in series_by_name:
            sid = len(series_order) + 1
            s = {
                "id": sid,
                "tmdb_id": None,
                "series": name,
                "year": None,
                "poster_url": None,
                "note": None,
                "needs_review": 1,
            }
            series_by_name[name] = s
            series_order.append(s)
        sid = series_by_name[name]["id"]
        season_no = row[hm["Season"]]
        owned_set = {
            TV_SERVICE[col]
            for col in TV_SERVICE
            if TV_SERVICE[col] and col in hm and is_yes(row[hm[col]])
        }
        seasons.append({
            "series_id": sid,
            "season": int(season_no) if isinstance(season_no, (int, float)) else season_no,
            "episode_count": None,
            "episodes": parse_episodes(row[hm["Episodes"]]),
            "owned_on": [s for s in TV_ORDER if s in owned_set],
        })

    os.makedirs(OUT, exist_ok=True)
    with open(os.path.join(OUT, "movies.json"), "w") as f:
        json.dump(movies, f, indent=2)
    with open(os.path.join(OUT, "tv.json"), "w") as f:
        json.dump({"series": series_order, "seasons": seasons}, f, indent=2)

    print(f"Imported {len(movies)} movies; "
          f"{len(series_order)} TV series / {len(seasons)} seasons. "
          f"(games.json untouched)")


if __name__ == "__main__":
    main()
